from pathlib import Path

from openpyxl import Workbook

from main import ScreenshotApp
from utils.course_db import CourseRepository, NewCourseSession
from utils.excel_import import ExcelCourseImporter
from utils.course_selection import CourseSessionAdapter


class DummyLogger:
    def __init__(self):
        self.messages = []

    def log(self, message):
        self.messages.append(message)


class DummyTree:
    def __init__(self):
        self.rows = {}

    def get_children(self):
        return list(self.rows.keys())

    def delete(self, item_id):
        self.rows.pop(item_id, None)

    def insert(self, parent, index, iid, values):
        self.rows[iid] = values


def test_course_repository_initializes_and_lists_sessions(tmp_path: Path):
    repo = CourseRepository(tmp_path / "optimad.db")

    inserted = repo.replace_all_sessions(
        [
            NewCourseSession(
                course_date="2026-04-06",
                start_time="18:00",
                end_time="21:00",
                duration_hours=3.0,
                course_name="Excel Avansat",
                group_name="Grupa A",
                trainer="Ion Popescu",
                is_recurring="DA",
                weekdays="LUNI",
                platform="zoom",
                meeting_link="https://zoom.us/test",
            )
        ]
    )

    assert inserted == 1
    sessions = repo.list_sessions()
    assert len(sessions) == 1
    assert sessions[0].course_name == "Excel Avansat"
    assert sessions[0].is_recurring == "DA"
    assert sessions[0].weekdays == "LUNI"
    assert repo.count_sessions() == 1


def test_course_repository_replace_all_and_delete(tmp_path: Path):
    repo = CourseRepository(tmp_path / "optimad.db")
    repo.replace_all_sessions(
        [
            NewCourseSession(
                course_date="2026-04-06",
                start_time="18:00",
                end_time="20:00",
                duration_hours=2.0,
                course_name="Primul curs",
            )
        ]
    )

    repo.replace_all_sessions(
        [
            NewCourseSession(
                course_date="2026-04-07",
                start_time="10:00",
                end_time="12:00",
                duration_hours=2.0,
                course_name="Al doilea curs",
            )
        ]
    )
    sessions = repo.list_sessions()
    assert len(sessions) == 1
    assert sessions[0].course_name == "Al doilea curs"

    repo.delete_session(sessions[0].id)
    assert repo.count_sessions() == 0


def test_course_repository_clear_all(tmp_path: Path):
    repo = CourseRepository(tmp_path / "optimad.db")
    repo.replace_all_sessions(
        [
            NewCourseSession(
                course_date="2026-04-06",
                start_time="18:00",
                end_time="20:00",
                duration_hours=2.0,
                course_name="Curs test",
            )
        ]
    )

    repo.clear_all_sessions()
    assert repo.list_sessions() == []


def test_excel_importer_parses_desktop_format_and_maps_links(tmp_path: Path):
    workbook = Workbook()
    course_sheet = workbook.active
    course_sheet.title = "Grafic cursuri"
    course_sheet.append(["Data", "Interval", "Curs", "Grupa", "Trainer", "Recurent", "Zile"])
    course_sheet.append(
        ["27.01.2026", "18:00-21:00", "Excel Avansat", "Grupa A", "Ion", "DA", "MARTI"]
    )

    link_sheet = workbook.create_sheet("Link-uri zoom")
    link_sheet.append(["Curs", "LUNI", "MARTI"])
    link_sheet.append(
        [
            "Excel Avansat",
            "",
            "Join Zoom Meeting https://zoom.us/excel-avansat Meeting ID: 123 Passcode: abc",
        ]
    )

    file_path = tmp_path / "import.xlsx"
    workbook.save(file_path)

    result = ExcelCourseImporter().import_file(file_path)

    assert result.errors == []
    assert result.imported_count == 1
    assert result.skipped_count == 0
    assert result.sessions[0].course_date == "2026-01-27"
    assert result.sessions[0].start_time == "18:00"
    assert result.sessions[0].end_time == "21:00"
    assert result.sessions[0].duration_hours == 3.0
    assert result.sessions[0].meeting_link == "https://zoom.us/excel-avansat"
    assert result.sessions[0].platform == "zoom"
    assert result.sessions[0].is_recurring == "DA"
    assert result.sessions[0].weekdays == "MARTI"


def test_excel_importer_skips_malformed_rows(tmp_path: Path):
    workbook = Workbook()
    course_sheet = workbook.active
    course_sheet.title = "Grafic cursuri"
    course_sheet.append(["Data", "Interval", "Curs"])
    course_sheet.append(["27.01.2026", "bad-interval", "Excel"])
    course_sheet.append([None, "18:00-19:00", "Word"])

    file_path = tmp_path / "invalid.xlsx"
    workbook.save(file_path)

    result = ExcelCourseImporter().import_file(file_path)

    assert result.imported_count == 0
    assert result.skipped_count == 2
    assert len(result.warnings) == 3
    assert "Link-uri zoom" in result.warnings[0]


def test_refresh_course_sessions_updates_treeview(tmp_path: Path):
    repo = CourseRepository(tmp_path / "optimad.db")
    repo.replace_all_sessions(
        [
            NewCourseSession(
                course_date="2026-04-06",
                start_time="18:00",
                end_time="20:00",
                duration_hours=2.0,
                course_name="Excel Avansat",
                group_name="Grupa A",
                trainer="Ion",
                is_recurring="DA",
                weekdays="LUNI",
                platform="zoom",
                meeting_link="https://zoom.us/test",
            )
        ]
    )

    app = ScreenshotApp.__new__(ScreenshotApp)
    app.course_repository = repo
    app.course_tree = DummyTree()
    app.logger = DummyLogger()
    app.course_adapter = CourseSessionAdapter()

    app._refresh_course_sessions()

    assert len(app.course_tree.rows) == 1
    row = next(iter(app.course_tree.rows.values()))
    assert row[0] == "2026-04-06"
    assert row[4] == "Excel Avansat"
    assert row[7] == "DA"
    assert row[8] == "LUNI"
    assert row[9] == "zoom"


def test_format_import_result_includes_warnings():
    app = ScreenshotApp.__new__(ScreenshotApp)
    result = type(
        "Result",
        (),
        {
            "warnings": ["Randul 2: interval invalid", "Randul 3: data lipsa"],
            "skipped_count": 2,
        },
    )()

    message = app._format_import_result(result, inserted=4)

    assert "Sesiuni importate: 4." in message
    assert "Randuri sarite: 2." in message
    assert "- Randul 2: interval invalid" in message
