# Script pentru verificarea numerelor de telefon pe portabilitate.ro
# Autor: Sima Alexandru
# Descriere: Preia numere de telefon dintr-un Excel, verifica operatorul curent
#            prin scraping portabilitate.ro si exporta rezultatele.

import os
import time
import sys
from datetime import datetime
from typing import Optional

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Configurari ──────────────────────────────────────────────────────────────
PORTABILITATE_URL = "https://www.portabilitate.ro/verifica"
REQUEST_DELAY = 1.5        # secunde intre cereri (respecta rate limiting)
REQUEST_TIMEOUT = 15       # timeout per cerere HTTP
MAX_RETRIES = 3            # reincercari per numar in caz de eroare
BATCH_SIZE = 50            # numarul de randuri procesate inainte de salvare partiala

INPUT_FILE = "numere_verificare.xlsx"   # fisierul cu numerele de verificat
OUTPUT_DIR = "output_portabilitate"     # directorul de iesire
OUTPUT_PREFIX = "Rezultate_Portabilitate"

# Coloana din fisierul de intrare care contine numerele
PHONE_COLUMN = "Telefon"

# Operatori mobili Romania
OPERATORI_CUNOSCUTI = [
    "Orange", "Vodafone", "Digi", "Telekom", "RCS&RDS",
    "Lycamobile", "Cosmote", "Zapp"
]

HEADER_FILL = "1F4E79"
HEADER_FONT = "FFFFFF"
SUCCESS_FILL = "C6EFCE"
ERROR_FILL = "FFC7CE"
ALT_FILL = "EBF3FB"


def ensure_output_dir():
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def normalize_phone(phone: str) -> str:
    """
    Normalizeaza un numar de telefon:
    - Sterge spatii, liniute, paranteze
    - Adauga prefixul +40 daca lipseste
    """
    cleaned = "".join(filter(str.isdigit, str(phone)))
    if cleaned.startswith("40") and len(cleaned) == 11:
        return "+" + cleaned
    if cleaned.startswith("0") and len(cleaned) == 10:
        return "+4" + cleaned
    if len(cleaned) == 9:
        return "+40" + cleaned
    return cleaned


def check_portability(phone: str, session: requests.Session) -> dict:
    """
    Verifica operatorul unui numar pe portabilitate.ro.
    Returneaza un dict cu: numar, operator, portabil, eroare.
    """
    result = {
        "Numar_Original": phone,
        "Numar_Normalizat": "",
        "Operator": "",
        "Portabil": "",
        "Data_Verificare": datetime.now().strftime("%d.%m.%Y %H:%M"),
        "Status": "",
        "Eroare": "",
    }

    normalized = normalize_phone(phone)
    result["Numar_Normalizat"] = normalized

    if len("".join(filter(str.isdigit, normalized))) < 9:
        result["Status"] = "INVALID"
        result["Eroare"] = "Numar prea scurt"
        return result

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = session.post(
                PORTABILITATE_URL,
                data={"phone": normalized},
                timeout=REQUEST_TIMEOUT,
                headers={
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                                  "Chrome/120.0.0.0 Safari/537.36",
                    "Accept-Language": "ro-RO,ro;q=0.9",
                    "Referer": "https://www.portabilitate.ro/",
                }
            )
            response.raise_for_status()

            soup = BeautifulSoup(response.text, "html.parser")

            # Cauta rezultatul in pagina
            result_div = soup.find("div", {"class": "result"}) or \
                         soup.find("div", {"id": "result"}) or \
                         soup.find("p", {"class": "operator"})

            if result_div:
                text = result_div.get_text(strip=True)
                result["Operator"] = text
                result["Status"] = "OK"

                # Detecteaza daca e portabil
                portabil_div = soup.find(text=lambda t: t and "portab" in t.lower())
                result["Portabil"] = "Da" if portabil_div else "Nu"
            else:
                # Fallback: cauta orice operator cunoscut in pagina
                page_text = soup.get_text()
                found_operator = next(
                    (op for op in OPERATORI_CUNOSCUTI if op.lower() in page_text.lower()),
                    None
                )
                if found_operator:
                    result["Operator"] = found_operator
                    result["Status"] = "OK"
                else:
                    result["Status"] = "NECUNOSCUT"
                    result["Eroare"] = "Operator negasit in raspuns"

            break  # iesire din bucla retry daca a reusit

        except requests.exceptions.Timeout:
            result["Eroare"] = f"Timeout (incercare {attempt}/{MAX_RETRIES})"
            if attempt < MAX_RETRIES:
                time.sleep(REQUEST_DELAY * attempt)
        except requests.exceptions.RequestException as e:
            result["Eroare"] = f"Eroare retea: {str(e)[:80]}"
            if attempt < MAX_RETRIES:
                time.sleep(REQUEST_DELAY * attempt)
        except Exception as e:
            result["Eroare"] = f"Eroare neasteptata: {str(e)[:80]}"
            result["Status"] = "EROARE"
            break

    if not result["Status"]:
        result["Status"] = "EROARE"

    return result


def load_phones(input_file: str) -> list:
    """Incarca numerele de telefon din fisierul Excel."""
    if not os.path.exists(input_file):
        print(f"[EROARE] Fisierul '{input_file}' nu a fost gasit.")
        sys.exit(1)

    df = pd.read_excel(input_file, dtype=str)

    if PHONE_COLUMN not in df.columns:
        print(f"[EROARE] Coloana '{PHONE_COLUMN}' nu exista in fisier.")
        print(f"[INFO] Coloane disponibile: {list(df.columns)}")
        sys.exit(1)

    phones = df[PHONE_COLUMN].dropna().unique().tolist()
    print(f"[INFO] {len(phones)} numere unice de verificat.")
    return phones


def style_results_sheet(ws, df: pd.DataFrame, title: str):
    """Aplica stiluri pe sheet-ul de rezultate."""
    header_font = Font(name="Calibri", bold=True, color=HEADER_FONT, size=11)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Titlu
    ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
    t = ws["A1"]
    t.value = title
    t.font = Font(name="Calibri", bold=True, size=14, color=HEADER_FONT)
    t.fill = header_fill
    t.alignment = center
    ws.row_dimensions[1].height = 26

    # Header
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=2, column=ci, value=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin

    # Date cu culoare pe status
    for ri, row in enumerate(df.itertuples(index=False), 3):
        status = str(getattr(row, "Status", ""))
        if status == "OK":
            row_fill = PatternFill("solid", fgColor=SUCCESS_FILL)
        elif status in ("EROARE", "INVALID"):
            row_fill = PatternFill("solid", fgColor=ERROR_FILL)
        else:
            row_fill = PatternFill("solid", fgColor=ALT_FILL) if ri % 2 == 0 else PatternFill()

        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = row_fill
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = thin

    for ci, col in enumerate(df.columns, 1):
        max_len = max(
            len(str(col)),
            df.iloc[:, ci - 1].astype(str).str.len().max() if not df.empty else 0
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 45)


def export_results(results: list):
    """Exporta rezultatele verificarii intr-un fisier Excel."""
    df = pd.DataFrame(results)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{OUTPUT_PREFIX}_{timestamp}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = Workbook()

    # Sheet 1: Toate rezultatele
    ws_all = wb.active
    ws_all.title = "Toate Rezultatele"
    style_results_sheet(ws_all, df, f"Verificare Portabilitate - {datetime.now().strftime('%d.%m.%Y')}")

    # Sheet 2: Doar erorile
    df_errors = df[df["Status"] != "OK"]
    if not df_errors.empty:
        ws_err = wb.create_sheet(title="Erori")
        style_results_sheet(ws_err, df_errors.reset_index(drop=True), "Numere cu Erori")

    # Sheet 3: Statistici per operator
    if "Operator" in df.columns:
        op_stats = (
            df[df["Operator"] != ""]
            .groupby("Operator")
            .size()
            .reset_index(name="Numar_Abonati")
            .sort_values("Numar_Abonati", ascending=False)
        )
        if not op_stats.empty:
            ws_stats = wb.create_sheet(title="Statistici Operatori")
            style_results_sheet(ws_stats, op_stats, "Distributie pe Operatori")

    wb.save(filepath)
    print(f"[SUCCES] Rezultate exportate: {filepath}")
    print(f"[INFO] Total: {len(df)} | OK: {len(df[df['Status']=='OK'])} | Erori: {len(df[df['Status']!='OK'])}")
    return filepath


def run(input_file: str = INPUT_FILE):
    """
    Punctul principal de intrare.
    input_file: calea catre fisierul Excel cu numerele de verificat.
    """
    print("=" * 60)
    print(f"  Verificare Portabilitate - {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    print(f"  Sursa: {input_file}")
    print("=" * 60)

    ensure_output_dir()
    phones = load_phones(input_file)

    results = []
    session = requests.Session()

    for idx, phone in enumerate(phones, 1):
        print(f"  [{idx}/{len(phones)}] Verific: {phone} ...", end=" ")
        result = check_portability(phone, session)
        results.append(result)
        print(f"{result['Status']} | {result.get('Operator', '')}")

        # Salvare partiala la fiecare BATCH_SIZE numere
        if idx % BATCH_SIZE == 0:
            print(f"  [CHECKPOINT] Salvare partiala la {idx} numere...")
            export_results(results)

        time.sleep(REQUEST_DELAY)

    export_results(results)
    session.close()
    print("[FINALIZAT] Verificare completa.")
    print("=" * 60)


if __name__ == "__main__":
    input_path = sys.argv[1] if len(sys.argv) > 1 else INPUT_FILE
    run(input_file=input_path)
