# Script pentru generarea rapoartelor BCR Lunare
# Autor: Sima Alexandru
# Descriere: Proceseaza fisiere Excel BCR si genereaza rapoarte lunare consolidate

import os
import re
import glob
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Configurari ──────────────────────────────────────────────────────────────
INPUT_DIR = "input_bcr"        # folderul cu fisierele Excel de intrare
OUTPUT_DIR = "output_rapoarte" # folderul unde se salveaza rapoartele generate
REPORT_PREFIX = "Raport_BCR"   # prefixul numelui fisierului de iesire

# Coloane asteptate in fisierul sursa BCR
REQUIRED_COLUMNS = [
    "Nume", "Prenume", "Telefon", "Email",
    "Produs", "Status", "Data", "Agent"
]

# Culori header tabel
HEADER_FILL_COLOR = "1F4E79"  # albastru inchis BCR
HEADER_FONT_COLOR = "FFFFFF"  # alb
ALT_ROW_COLOR = "D6E4F0"      # albastru deschis alternativ


def ensure_dirs():
    """Creeaza directoarele necesare daca nu exista."""
    os.makedirs(INPUT_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def load_excel_files(input_dir: str) -> pd.DataFrame:
    """Incarca si concateneaza toate fisierele Excel din directorul de intrare."""
    files = glob.glob(os.path.join(input_dir, "*.xlsx")) + \
            glob.glob(os.path.join(input_dir, "*.xls"))

    if not files:
        print(f"[AVERTISMENT] Nu au fost gasite fisiere Excel in '{input_dir}'.")
        return pd.DataFrame()

    frames = []
    for f in files:
        try:
            df = pd.read_excel(f, dtype=str)
            df["_sursa"] = os.path.basename(f)
            frames.append(df)
            print(f"[OK] Incarcat: {os.path.basename(f)} ({len(df)} randuri)")
        except Exception as e:
            print(f"[EROARE] Nu s-a putut incarca {f}: {e}")

    if not frames:
        return pd.DataFrame()

    combined = pd.concat(frames, ignore_index=True)
    return combined


def validate_columns(df: pd.DataFrame) -> bool:
    """Verifica daca coloanele necesare exista in DataFrame."""
    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        print(f"[EROARE] Coloane lipsa: {missing}")
        print(f"[INFO] Coloane disponibile: {list(df.columns)}")
        return False
    return True


def normalize_data(df: pd.DataFrame) -> pd.DataFrame:
    """Curata si normalizeaza datele."""
    df = df.copy()

    # Normalizeaza coloana Data
    if "Data" in df.columns:
        df["Data"] = pd.to_datetime(df["Data"], dayfirst=True, errors="coerce")
        df["Luna"] = df["Data"].dt.to_period("M").astype(str)
    else:
        df["Luna"] = "Necunoscut"

    # Normalizeaza telefon: sterge spatii si caractere non-numerice
    if "Telefon" in df.columns:
        df["Telefon"] = df["Telefon"].str.replace(r"\D", "", regex=True)

    # Trim whitespace pentru toate coloanele string
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].str.strip()

    return df


def generate_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Genereaza un rezumat agregat pe luna, produs si status."""
    if df.empty:
        return pd.DataFrame()

    summary = (
        df.groupby(["Luna", "Produs", "Status", "Agent"])
        .size()
        .reset_index(name="Numar_Cazuri")
        .sort_values(["Luna", "Produs", "Agent"])
    )
    return summary


def style_worksheet(ws, df: pd.DataFrame, sheet_title: str):
    """Aplica stiluri profesionale pe un worksheet openpyxl."""
    header_font = Font(name="Calibri", bold=True, color=HEADER_FONT_COLOR, size=11)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    alt_fill = PatternFill("solid", fgColor=ALT_ROW_COLOR)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Titlu raport
    ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
    title_cell = ws["A1"]
    title_cell.value = sheet_title
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=HEADER_FONT_COLOR)
    title_cell.fill = header_fill
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 28

    # Header coloane (rand 2)
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Date
    for row_idx, row in enumerate(df.itertuples(index=False), start=3):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border

    # Auto-width coloane
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col_name)),
            df.iloc[:, col_idx - 1].astype(str).str.len().max() if not df.empty else 0
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)


def export_report(df_data: pd.DataFrame, df_summary: pd.DataFrame, luna: Optional[str] = None):
    """Exporta raportul final intr-un fisier Excel cu mai multe sheet-uri."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    period_label = luna if luna else "Toate_Lunile"
    filename = f"{REPORT_PREFIX}_{period_label}_{timestamp}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = Workbook()

    # Sheet 1: Date complete
    ws_data = wb.active
    ws_data.title = "Date Complete"
    style_worksheet(ws_data, df_data, f"BCR - Date Complete ({period_label})")

    # Sheet 2: Rezumat
    if not df_summary.empty:
        ws_summary = wb.create_sheet(title="Rezumat")
        style_worksheet(ws_summary, df_summary, f"BCR - Rezumat Lunar ({period_label})")

    # Sheet 3: Statistici per Agent
    if "Agent" in df_data.columns and "Numar_Cazuri" in df_summary.columns:
        agent_stats = (
            df_summary.groupby("Agent")["Numar_Cazuri"]
            .sum()
            .reset_index()
            .sort_values("Numar_Cazuri", ascending=False)
        )
        ws_agents = wb.create_sheet(title="Statistici Agenti")
        style_worksheet(ws_agents, agent_stats, "BCR - Performanta Agenti")

    wb.save(filepath)
    print(f"[SUCCES] Raport generat: {filepath}")
    return filepath


def run(luna_filter: Optional[str] = None):
    """
    Punctul principal de intrare.
    luna_filter: ex. '2025-01' pentru a filtra doar o luna specifica.
                 None = proceseaza toate lunile disponibile.
    """
    print("=" * 60)
    print(f"  Generare Rapoarte BCR Lunare - {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    print("=" * 60)

    ensure_dirs()

    df = load_excel_files(INPUT_DIR)
    if df.empty:
        print("[INFO] Nu exista date de procesat.")
        return

    if not validate_columns(df):
        print("[INFO] Verificati structura fisierelor de intrare.")
        return

    df = normalize_data(df)

    # Filtrare optionala pe luna
    if luna_filter:
        df = df[df["Luna"] == luna_filter]
        if df.empty:
            print(f"[AVERTISMENT] Nu exista date pentru luna: {luna_filter}")
            return
        print(f"[INFO] Filtrat pentru luna: {luna_filter} ({len(df)} randuri)")
    else:
        print(f"[INFO] Total randuri procesate: {len(df)}")

    summary = generate_summary(df)

    # Elimina coloana interna _sursa inainte de export
    export_df = df.drop(columns=["_sursa", "Luna"], errors="ignore")

    export_report(export_df, summary, luna=luna_filter)
    print("[FINALIZAT] Procesare completa.")
    print("=" * 60)


if __name__ == "__main__":
    import sys
    luna_arg = sys.argv[1] if len(sys.argv) > 1 else None
    run(luna_filter=luna_arg)
