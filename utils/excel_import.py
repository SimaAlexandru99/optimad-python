from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
import re
from typing import Dict, List, Sequence

from openpyxl import load_workbook

from .course_db import NewCourseSession


@dataclass
class ImportResult:
    sessions: List[NewCourseSession] = field(default_factory=list)
    imported_count: int = 0
    skipped_count: int = 0
    warnings: List[str] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)


class ExcelCourseImporter:
    """Parse the desktop Excel schedule format into local course sessions."""

    COURSE_SHEET = "Grafic cursuri"
    LINK_SHEET = "Link-uri zoom"
    URL_PATTERN = re.compile(r"https?://\S+")

    def import_file(self, file_path: str | Path) -> ImportResult:
        workbook = load_workbook(filename=Path(file_path), data_only=True)
        result = ImportResult()

        if self.COURSE_SHEET not in workbook.sheetnames:
            result.errors.append(
                f"Lipseste sheet-ul obligatoriu '{self.COURSE_SHEET}'."
            )
            return result

        link_mapping = {}
        if self.LINK_SHEET in workbook.sheetnames:
            link_mapping = self._build_link_mapping(workbook[self.LINK_SHEET])
        else:
            result.warnings.append(
                f"Sheet-ul optional '{self.LINK_SHEET}' nu exista. Linkurile nu au fost mapate."
            )

        course_sheet = workbook[self.COURSE_SHEET]
        course_headers = self._header_map(next(course_sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        for row_index, row in enumerate(
            course_sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            try:
                session = self._parse_course_row(row, course_headers, link_mapping)
            except ValueError as exc:
                result.skipped_count += 1
                result.warnings.append(f"Randul {row_index}: {exc}")
                continue

            result.sessions.append(session)

        result.imported_count = len(result.sessions)
        return result

    def _build_link_mapping(self, sheet) -> Dict[tuple[str, str], str]:
        mapping: Dict[tuple[str, str], str] = {}
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [self._normalize_text(value).upper() for value in header_row]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            course_name = self._normalize_text(row[0])
            for index, value in enumerate(row[1:], start=1):
                url = self._extract_url(value)
                if url:
                    weekday = headers[index] if index < len(headers) else ""
                    mapping[(course_name, weekday)] = url
        return mapping

    def _parse_course_row(
        self,
        row: Sequence[object],
        headers: Dict[str, int],
        link_mapping: Dict[tuple[str, str], str],
    ) -> NewCourseSession:
        if len(row) < 3:
            raise ValueError("Prea putine coloane pentru un curs valid.")

        course_date = self._parse_date(self._get_value(row, headers, "DATA"))
        start_time, end_time, duration_hours = self._parse_time_range(
            self._get_value(row, headers, "INTERVAL")
        )
        course_name = self._normalize_required_text(
            self._get_value(row, headers, "CURS"),
            "Numele cursului lipseste.",
        )
        group_name = self._normalize_text(self._get_value(row, headers, "GRUPA"))
        trainer = self._normalize_text(self._get_value(row, headers, "TRAINER"))
        is_recurring = self._normalize_text(self._get_value(row, headers, "RECURENT")).upper()
        weekdays = self._normalize_text(self._get_value(row, headers, "ZILE")).upper()

        meeting_link = (
            link_mapping.get((self._normalize_text(course_name), weekdays), "")
            or link_mapping.get((self._normalize_text(course_name), ""), "")
        )
        platform = self._infer_platform(meeting_link)

        return NewCourseSession(
            course_date=course_date,
            start_time=start_time,
            end_time=end_time,
            duration_hours=duration_hours,
            course_name=course_name,
            group_name=group_name,
            trainer=trainer,
            is_recurring=is_recurring,
            weekdays=weekdays,
            platform=platform,
            meeting_link=meeting_link,
        )

    def _parse_date(self, value: object) -> str:
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")

        text = self._normalize_required_text(value, "Data cursului lipseste.")
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        raise ValueError(f"Data invalida: '{text}'.")

    def _parse_time_range(self, value: object) -> tuple[str, str, float]:
        text = self._normalize_required_text(value, "Intervalul orar lipseste.")
        if "-" not in text:
            raise ValueError(f"Interval orar invalid: '{text}'.")

        start_raw, end_raw = [part.strip() for part in text.split("-", 1)]
        start_time = datetime.strptime(start_raw, "%H:%M")
        end_time = datetime.strptime(end_raw, "%H:%M")
        duration = (end_time - start_time).total_seconds() / 3600
        if duration <= 0:
            raise ValueError(f"Interval orar invalid: '{text}'.")

        return start_time.strftime("%H:%M"), end_time.strftime("%H:%M"), duration

    def _normalize_required_text(self, value: object, error_message: str) -> str:
        text = self._normalize_text(value)
        if not text:
            raise ValueError(error_message)
        return text

    def _normalize_text(self, value: object) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _header_map(self, header_row: Sequence[object]) -> Dict[str, int]:
        return {
            self._normalize_text(value).upper(): index
            for index, value in enumerate(header_row)
            if self._normalize_text(value)
        }

    def _get_value(
        self, row: Sequence[object], headers: Dict[str, int], header_name: str
    ) -> object:
        index = headers.get(header_name)
        if index is None or index >= len(row):
            return None
        return row[index]

    def _extract_url(self, value: object) -> str:
        text = self._normalize_text(value)
        if not text:
            return ""
        match = self.URL_PATTERN.search(text)
        return match.group(0) if match else ""

    def _infer_platform(self, meeting_link: str) -> str:
        link = meeting_link.lower()
        if "zoom.us" in link:
            return "zoom"
        if "teams." in link or "microsoft.com" in link:
            return "teams"
        return ""
